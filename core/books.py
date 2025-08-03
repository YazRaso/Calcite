from openpyxl import load_workbook
from core import receipt


class ExcelManager:
    """
    A class representing an Excel workbook.
    Attributes:
        filepath (str): Path to the Excel workbook
    Methods:
        load(): Creates a new Excel workbook if it doesn't exist else loads it.
        add_transaction(
        amount, currency, conversion_rate, transaction_date, reference_id
        ): Adds a transaction to the Excel workbook.
        delete_transaction(): Deletes a transaction from the Excel workbook.
    """

    def __init__(self, filepath: str) -> None:
        """
        Initializes the Excel workbook, with a given filepath.
        :param filepath (str): Path to the Excel workbook:
        """
        super().__init__()
        self.filepath = filepath
        self.headers = ['Amount', 'Currency',
                        'Conversion Rate', 'Transaction Date',
                        'Reference ID']

        self.load()

    def load(self) -> None:
        """
        Loads a given Excel workbook
        :return: None
        """
        self.wb = load_workbook(self.filepath)
        self.ws = self.wb.active
        for col_num, header in enumerate(self.headers, start=1):
            self.ws.cell(row=1, column=col_num, value=header)

        self.save()

    def add_transaction(self, amount: float, currency: str, conversion_rate: float,
                        transaction_date: str, reference_id: str) -> None:
        """
        Adds a transaction to the last row of Excel workbook.
        :param amount:
        :param currency:
        :param conversion_rate:
        :param transaction_date:
        :param reference_id:
        :return: None
        """
        row = [
            amount,
            currency,
            conversion_rate,
            transaction_date,
            reference_id
        ]

        self.ws.append(row)
        self.save()

    def delete_transaction(self, reference_id: str) -> bool:
        first_row_number = 2
        number_of_cols = 5
        reference_id_column = 4
        for row in self.ws.iter_rows(
                min_row=first_row_number,
                max_col=number_of_cols):
            # Get transaction ID
            if row[reference_id_column].value == reference_id:
                self.ws.delete_rows(row[0].row)
                self.save()
                return True
        return False

    def generate_receipt_by_id(self, received_by, reference_id):
        first_row_number = 2
        number_of_cols = 5
        reference_id_column = 4
        for row in self.ws.iter_rows(min_row=first_row_number,
                                     max_col=number_of_cols):
            if row[reference_id_column].value == reference_id:
                row_number = row[0].row  # Get the actual row number
                amount = self.ws.cell(row=row_number, column=1).value
                currency = self.ws.cell(row=row_number, column=2).value
                transaction_date = self.ws.cell(row=row_number, column=4).value
                reference_id = self.ws.cell(row=row_number, column=5).value
                receipt_name = receipt.generate_receipt(received_by=str(received_by), reference_id=str(reference_id), amount=amount, currency=str(currency),
                                                        transaction_date=str(transaction_date))
                return receipt_name
        return "No such transaction was found, no receipt generated"

    def generate_receipt(self, received_by):
        last_row = self.ws.max_row  # 1 based index of latest transaction
        if last_row <= 1:
            return
        amount = self.ws.cell(row=last_row, column=1).value
        currency = self.ws.cell(row=last_row, column=2).value
        transaction_date = self.ws.cell(row=last_row, column=4).value
        reference_id = self.ws.cell(row=last_row, column=5).value
        receipt_name = receipt.generate_receipt(received_by=str(received_by), reference_id=str(reference_id), amount=amount, currency=str(currency),
                                                transaction_date=str(transaction_date))
        return receipt_name

    def save(self) -> None:
        self.wb.save(self.filepath)
