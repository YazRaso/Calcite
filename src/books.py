from openpyxl import Workbook, load_workbook
from pathlib import Path
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
import datetime


class ExcelManager:
    def __init__(self, filepath: str) -> None:
        self.filepath = Path(filepath)
        self.headers = ['Date', 'Amount', 'Currency',
                        'Conversion Rate', 'Transaction ID',
                        'Transaction Date']
        self.load_or_create()

    def load_or_create(self) -> None:
        if self.filepath.exists():
            self.wb = load_workbook(self.filepath)
            self.ws = self.wb.active
        else:
            self.wb = Workbook()
            self.ws = self.wb.active
            self.ws.append(self.headers)
            self.save()

    def add_transaction(self, amount: float, currency: str, conversion_rate: float,
                        transaction_id: str, transaction_date: str) -> None:
        row = [
            datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            amount,
            currency,
            conversion_rate,
            transaction_id,
            transaction_date
        ]
        self.ws.append(row)
        self.save()

    def delete_transaction(self, transaction_id: str) -> bool:
        for row in self.ws.iter_rows(min_row=2, max_col=6):
            if row[4].value == transaction_id:  # Match the transaction ID to delete
                self.ws.delete_rows(row[0].row)
                self.save()
                print(f"Transaction with ID '{transaction_id}' deleted.")
                return True
        print(f"No transaction found with ID '{transaction_id}'.")
        return False

    def save(self) -> None:
        self.wb.save(self.filepath)

    # TODO: Technical Debt: Receipt not formatted
    def generate_receipt(self, transaction_id) -> None:
        filename = f"{transaction_id} receipt.pdf"
        c = canvas.Canvas(filename, pagesize=A4)
        width, height = A4

        # Draw a horizontal line near the top
        c.line(50, height - 100, width - 50, height - 100)

        # Draw a vertical line on the left side
        c.line(50, height - 100, 50, 100)

        # Draw a horizontal line near the bottom
        c.line(50, 100, width - 50, 100)

        # Draw another vertical line on the right side
        c.line(width - 50, height - 100, width - 50, 100)

        # Optional: Title or label
        c.drawString(60, height - 80, f"Receipt ID: {transaction_id}")

        # Finalize and save the PDF
        c.showPage()
        c.save()



