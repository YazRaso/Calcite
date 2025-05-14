import unittest
import os
from openpyxl import load_workbook
from books import ExcelManager  # Assuming your ExcelManager class is in ExcelManager.py


class TestExcelManager(unittest.TestCase):
    def setUp(self):
        # Create a temporary Excel file for testing
        self.test_file = 'test_transactions.xlsx'
        self.manager = ExcelManager(self.test_file)

    def tearDown(self):
        # Clean up the test file after each test
        if os.path.exists(self.test_file):
            os.remove(self.test_file)

    def test_add_transaction(self):
        # Add a transaction and check if it appears in the file
        self.manager.add_transaction(100.5, 'USD', 1.1, 'T123', '2025-05-13')
        # Load the workbook and check if the row has been added
        wb = load_workbook(self.test_file)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, max_col=6, values_only=True))
        self.assertEqual(len(rows), 1)  # One row should be added
        self.assertEqual(rows[0][1], 100.5)  # Check the amount
        self.assertEqual(rows[0][2], 'USD')  # Check the currency
        self.assertEqual(rows[0][3], 1.1)  # Check the conversion rate
        self.assertEqual(rows[0][4], 'T123')  # Check the transaction ID
        self.assertEqual(rows[0][5], '2025-05-13')  # Check the transaction date

    def test_delete_transaction(self):
        # Add a transaction and then delete it
        self.manager.add_transaction(100.5, 'USD', 1.1, 'T123', '2025-05-13')
        # Delete the transaction
        deleted = self.manager.delete_transaction('T123')
        self.assertTrue(deleted)  # Check that deletion was successful

        # Verify that the file is empty after deletion
        wb = load_workbook(self.test_file)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, max_col=6, values_only=True))
        self.assertEqual(len(rows), 0)  # No rows should remain

    def test_delete_non_existent_transaction(self):
        # Attempt to delete a transaction that doesn't exist
        deleted = self.manager.delete_transaction('T999')
        self.assertFalse(deleted)  # Deletion should fail

    def test_save(self):
        # Add a transaction and check if it's saved correctly
        self.manager.add_transaction(50.0, 'EUR', 1.2, 'T124', '2025-05-14')
        # Reopen the file and check the saved data
        wb = load_workbook(self.test_file)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, max_col=6, values_only=True))
        self.assertEqual(rows[0][1], 50.0)  # Check that the transaction is saved

    def test_make_receipt(self):
        self.manager.generate_receipt('T123')
        self.assertTrue(os.path.exists('T123 receipt.pdf'))


if __name__ == '__main__':
    # unittest.main()
    tester = TestExcelManager()
    tester.setUp()
    tester.test_add_transaction()
    tester.test_make_receipt()
